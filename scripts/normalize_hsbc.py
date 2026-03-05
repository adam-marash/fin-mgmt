"""Normalize HSBC Guernsey activity statement XLSX into a clean CSV."""

import csv
import sys
from pathlib import Path

import openpyxl


def normalize(input_path: str, output_path: str | None = None):
    wb = openpyxl.load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]

    if output_path is None:
        output_path = str(Path(input_path).with_suffix(".csv"))

    portfolio = None
    sub_account = None
    currency = None
    rows = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        cell0 = str(vals[0]).strip() if vals[0] else ""
        cell1 = str(vals[1]).strip() if vals[1] else ""
        cell3 = str(vals[3]).strip() if vals[3] else ""

        # Skip the column header row
        if cell1 == "Currency" and cell3 == "Value Date":
            continue

        # Skip report header lines
        if cell0.startswith(("Value Date:", "Bookings:", "RM:", "Mandate:")):
            continue

        # Client ID line (e.g. "B7RTW2Z *** -")
        if cell0 and "***" in cell0:
            continue

        # Portfolio header (e.g. "C42ITVI Standard INCOME GBP")
        if cell0 and not vals[1] and not vals[2] and not vals[3]:
            parts = cell0.split(None, 1)
            portfolio = parts[0] if parts else cell0
            continue

        # Sub-account + currency line (e.g. "M7VFNELG" in col 0, "EUR" in col 1)
        if cell0 and cell1 in ("EUR", "USD", "GBP", "ILS", "CHF"):
            sub_account = cell0
            currency = cell1
            continue

        # Opening/closing balance
        if "balance" in cell3.lower():
            rows.append({
                "portfolio": portfolio,
                "sub_account": sub_account,
                "currency": currency,
                "date": "",
                "value_date": "",
                "amount": vals[4],
                "description": cell3,
                "order_nr": "",
                "type": "balance",
            })
            continue

        # Transaction row: verification_date in col 2, value_date in col 3
        if vals[2] is not None and vals[4] is not None and cell3 and "balance" not in cell3.lower():
            verif_date = str(vals[2])[:10] if vals[2] else ""
            value_date = str(vals[3])[:10] if vals[3] else ""
            description = str(vals[5]) if vals[5] else ""
            order_nr = str(vals[6]) if vals[6] else ""

            # Parse description into type and detail
            tx_type = ""
            detail = description
            if " - " in description:
                tx_type, detail = description.split(" - ", 1)
                tx_type = tx_type.strip()
                detail = detail.strip()

            rows.append({
                "portfolio": portfolio,
                "sub_account": sub_account,
                "currency": currency,
                "date": verif_date,
                "value_date": value_date,
                "amount": vals[4],
                "description": detail,
                "order_nr": order_nr,
                "type": tx_type,
            })

    # Write CSV
    fieldnames = ["portfolio", "sub_account", "currency", "date", "value_date",
                   "amount", "type", "description", "order_nr"]
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} rows to {output_path}")
    # Summary
    txns = [r for r in rows if r["type"] != "balance"]
    print(f"  Transactions: {len(txns)}")
    print(f"  Balance entries: {len(rows) - len(txns)}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <input.xlsx> [output.csv]")
        sys.exit(1)
    output = sys.argv[2] if len(sys.argv) > 2 else None
    normalize(sys.argv[1], output)
