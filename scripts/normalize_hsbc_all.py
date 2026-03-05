"""Normalize HSBC Guernsey all-transactions XLSX into a clean CSV.

Reads the full-history export and produces:
  date, value_date, portfolio, sub_account, currency, amount, description, order_nr

Skips opening/closing balance rows.
"""

import csv
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


def normalize(input_path: str, output_path: str | None = None):
    wb = openpyxl.load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]

    if output_path is None:
        output_path = str(Path(input_path).with_suffix(".csv"))

    portfolio = None
    sub_account = None
    currency = None
    rows = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        cell0 = str(vals[0]).strip() if vals[0] else ""
        cell1 = str(vals[1]).strip() if vals[1] else ""
        cell3 = str(vals[3]).strip() if vals[3] else ""

        # Skip the column header row
        if cell1 == "Currency" and cell3 == "Value Date":
            continue

        # Skip report header lines
        if cell0.startswith(("Value Date:", "Bookings:", "RM:", "Mandate:")):
            continue

        # Client ID line (e.g. "B7RTW2Z *** -")
        if cell0 and "***" in cell0:
            continue

        # Portfolio header (e.g. "C42ITVI Standard INCOME GBP")
        if cell0 and not vals[1] and not vals[2] and not vals[3]:
            parts = cell0.split(None, 1)
            portfolio = parts[0] if parts else cell0
            continue

        # Sub-account + currency line (e.g. "M7VFNELG" in col 0, "EUR" in col 1)
        if cell0 and cell1 in ("EUR", "USD", "GBP", "ILS", "CHF", "PLN"):
            sub_account = cell0
            currency = cell1
            continue

        # Skip opening/closing balance rows
        if "balance" in cell3.lower():
            continue

        # Transaction row: verification_date in col 2, value_date in col 3
        if vals[2] is not None and vals[4] is not None:
            verif_date = vals[2]
            value_date = vals[3]

            # Format dates as YYYY-MM-DD
            if isinstance(verif_date, datetime):
                verif_date = verif_date.strftime("%Y-%m-%d")
            else:
                verif_date = str(verif_date)[:10]

            if isinstance(value_date, datetime):
                value_date = value_date.strftime("%Y-%m-%d")
            else:
                value_date = str(value_date)[:10]

            description = str(vals[5]) if vals[5] else ""
            order_nr = str(int(vals[6])) if vals[6] else ""

            rows.append({
                "date": verif_date,
                "value_date": value_date,
                "portfolio": portfolio,
                "sub_account": sub_account,
                "currency": currency,
                "amount": vals[4],
                "description": description,
                "order_nr": order_nr,
            })

    # Write CSV
    fieldnames = ["date", "value_date", "portfolio", "sub_account", "currency",
                   "amount", "description", "order_nr"]
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} rows to {output_path}")

    # Summary by portfolio/sub_account
    from collections import Counter
    combos = Counter((r["portfolio"], r["sub_account"], r["currency"]) for r in rows)
    for (p, s, c), n in sorted(combos.items()):
        print(f"  {p} / {s} ({c}): {n} txns")

    # Date range
    dates = sorted(r["date"] for r in rows if r["date"])
    if dates:
        print(f"  Date range: {dates[0]} to {dates[-1]}")

    return rows


def check_overlap(all_rows, existing_csv: str):
    """Check overlap between all-transactions and existing tamar-2025.csv."""
    import csv as csv_mod

    with open(existing_csv) as f:
        reader = csv_mod.DictReader(f)
        existing = []
        for r in reader:
            if r.get("type") == "balance":
                continue
            existing.append(r)

    # Build lookup by order_nr (unique transaction identifier)
    all_by_order = {r["order_nr"]: r for r in all_rows if r["order_nr"]}
    existing_by_order = {r["order_nr"]: r for r in existing if r.get("order_nr")}

    overlap = set(all_by_order.keys()) & set(existing_by_order.keys())
    only_existing = set(existing_by_order.keys()) - set(all_by_order.keys())
    only_all = set(all_by_order.keys()) - set(existing_by_order.keys())

    print(f"\nOverlap analysis vs {existing_csv}:")
    print(f"  Existing transactions (non-balance): {len(existing)}")
    print(f"  All-transactions rows: {len(all_rows)}")
    print(f"  Matching by order_nr: {len(overlap)}")
    print(f"  Only in existing: {len(only_existing)}")
    print(f"  Only in all-transactions: {len(only_all)}")

    if only_existing:
        print(f"\n  Order numbers only in existing (first 10):")
        for o in sorted(only_existing)[:10]:
            r = existing_by_order[o]
            print(f"    {o}: {r['date']} {r['amount']} {r['description'][:60]}")

    return len(overlap)


if __name__ == "__main__":
    input_path = "tmp/more/HSBC GU transactions - all.xlsx"
    output_path = "data/hsbc-gu/all-transactions.csv"

    if len(sys.argv) >= 2:
        input_path = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]

    rows = normalize(input_path, output_path)

    # Check overlap with existing file
    existing = Path("data/hsbc-gu/tamar-2025.csv")
    if existing.exists():
        check_overlap(rows, str(existing))
