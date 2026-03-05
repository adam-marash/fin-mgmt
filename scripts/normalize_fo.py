"""Normalize The Service (FO) transaction register XLSX into a clean CSV."""

import csv
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# Column indices (Hebrew headers)
COL_PARENT_CLIENT = 0   # לקוח אב
COL_CLIENT_MANAGER = 1  # מנהל לקוח
COL_OWNER = 2           # שייכות
COL_PRODUCT_TYPE = 3    # סוג מוצר
COL_INVESTMENT = 4      # תאור
COL_ACCOUNT_NUM = 5     # מספר חשבון
COL_MANAGING_BODY = 6   # גוף מנהל
COL_TX_TYPE = 7         # סוג תנועה
COL_EXT_TYPE = 8        # סוג תנועה מורחב
COL_DEPOSIT_TYPE = 9    # סוג הפקדה
COL_PROVISION_TYPE = 10  # סוג הפרשה
COL_DATE = 11           # תאריך התנועה
COL_AMOUNT = 12         # סכום תנועה במטבע
COL_CURRENCY = 13       # מטבע התנועה
COL_FX_RATE = 14        # שער המרה לתנועה
COL_AMOUNT_ILS = 15     # סכום תנועה בש"ח

# Transaction type translations
TX_TYPE_MAP = {
    "משיכת תשואה": "yield_withdrawal",
    "הפקדה": "deposit",
    "משיכה": "withdrawal",
}

EXT_TYPE_MAP = {
    "הפקדת הון ראשונית": "initial_capital",
    "הפקדה נוספת": "additional_deposit",
}

CURRENCY_MAP = {
    "$": "USD",
    "€": "EUR",
    "₪": "ILS",
    "£": "GBP",
}


def parse_date(val) -> tuple[str, bool]:
    """Returns (date_str, uncertain). Uncertain when Excel auto-parsed a datetime
    and we can't tell if DD/MM was swapped."""
    if not val:
        return ("", False)
    if isinstance(val, datetime):
        # Excel auto-parsed: if day is 1 (could be DD/MM swap artifact), flag it
        uncertain = val.day == 1 and val.month <= 12
        return (val.strftime("%Y-%m-%d"), uncertain)
    s = str(val).strip()
    # String dates in DD/MM/YYYY are unambiguous (FO format)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return (datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d"), False)
        except ValueError:
            continue
    return (s, False)


def normalize(input_path: str, output_path: str | None = None):
    wb = openpyxl.load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]

    if output_path is None:
        output_path = str(Path(input_path).with_suffix(".csv"))

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        investment = str(row[COL_INVESTMENT]) if row[COL_INVESTMENT] else ""
        if not investment:
            continue

        tx_type_heb = str(row[COL_TX_TYPE]) if row[COL_TX_TYPE] else ""
        ext_type_heb = str(row[COL_EXT_TYPE]) if row[COL_EXT_TYPE] else ""
        currency_raw = str(row[COL_CURRENCY]) if row[COL_CURRENCY] else ""

        date_str, date_uncertain = parse_date(row[COL_DATE])

        rows.append({
            "owner": str(row[COL_OWNER]) if row[COL_OWNER] else "",
            "investment": investment,
            "account_number": str(row[COL_ACCOUNT_NUM]).replace(".0", "") if row[COL_ACCOUNT_NUM] else "",
            "managing_body": str(row[COL_MANAGING_BODY]) if row[COL_MANAGING_BODY] else "",
            "tx_type": TX_TYPE_MAP.get(tx_type_heb, tx_type_heb),
            "ext_type": EXT_TYPE_MAP.get(ext_type_heb, ext_type_heb),
            "date": date_str,
            "date_uncertain": "Y" if date_uncertain else "",
            "amount": row[COL_AMOUNT] if row[COL_AMOUNT] is not None else "",
            "currency": CURRENCY_MAP.get(currency_raw, currency_raw),
            "fx_rate": row[COL_FX_RATE] if row[COL_FX_RATE] is not None else "",
            "amount_ils": row[COL_AMOUNT_ILS] if row[COL_AMOUNT_ILS] is not None else "",
        })

    # Sort by date
    rows.sort(key=lambda r: r["date"] or "9999")

    fieldnames = ["date", "date_uncertain", "owner", "investment", "account_number",
                   "managing_body", "tx_type", "ext_type", "amount", "currency",
                   "fx_rate", "amount_ils"]
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} rows to {output_path}")

    # Summary by investment
    from collections import Counter
    inv_counts = Counter(r["investment"] for r in rows)
    print(f"\nInvestments: {len(inv_counts)}")
    date_range = [r["date"] for r in rows if r["date"]]
    if date_range:
        print(f"Date range: {min(date_range)} to {max(date_range)}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <input.xlsx> [output.csv]")
        sys.exit(1)
    output = sys.argv[2] if len(sys.argv) > 2 else None
    normalize(sys.argv[1], output)
